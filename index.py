import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("transactions.xlsx")
sheet = wb['Sheet1']
# cell = sheet['a1']
# cell = sheet.cell(1, 1)
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_Value = cell.value * 0.9
    corrected_Value_cell = sheet.cell(row, 4)
    corrected_Value_cell.value = corrected_Value
    print(sheet.cell(row, 4).value)

values = Reference(sheet,min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'E2')

wb.save("transactions2.xlsx")
